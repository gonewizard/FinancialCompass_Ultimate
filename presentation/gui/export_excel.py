import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from tkinter import filedialog, messagebox
from core.entities.operation import OperationType


class ExcelExporter:
    def __init__(self, app, current_user):
        self.app = app
        self.current_user = current_user

    def export_operations(self, operations, filename=None):
        if not filename:
            filename = filedialog.asksaveasfilename(
                title="Экспорт операций в Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not filename:
                return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Операции"

        headers = ['Дата', 'Тип', 'Категория', 'Сумма', 'Контрагент', 'Описание']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row, op in enumerate(operations, 2):
            ws.cell(row=row, column=1, value=op.operation_date.strftime('%d.%m.%Y %H:%M'))
            ws.cell(row=row, column=2, value="Доход" if op.type.value == "income" else "Расход")
            ws.cell(row=row, column=3, value=op.category)
            ws.cell(row=row, column=4, value=float(op.amount))
            ws.cell(row=row, column=5, value=op.counterparty or "")
            ws.cell(row=row, column=6, value=op.description or "")

            for col in range(1, 7):
                ws.cell(row=row, column=col).border = thin_border

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['F'].width = 30

        wb.save(filename)
        messagebox.showinfo("Успех", f"Операции экспортированы в {filename}")
        return True

    def export_operations_filtered(self, filters=None, filename=None):
        operations = self.app.operation_service.get_user_operations(self.current_user.id)

        if filters:
            filtered = []
            for op in operations:
                if filters.get('category') and filters['category'].lower() not in op.category.lower():
                    continue
                if filters.get('counterparty') and (
                        not op.counterparty or filters['counterparty'].lower() not in op.counterparty.lower()):
                    continue
                if filters.get('type'):
                    op_type = 'income' if op.type == OperationType.INCOME else 'expense'
                    if op_type != filters['type']:
                        continue
                filtered.append(op)
            operations = filtered

        return self.export_operations(operations, filename)

    def export_category_stats(self, date_from=None, date_to=None, filename=None):
        if not filename:
            filename = filedialog.asksaveasfilename(
                title="Экспорт статистики по категориям",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not filename:
                return False

        operations = self.app.operation_service.get_user_operations(self.current_user.id)

        expense_data = {}
        income_data = {}

        for op in operations:
            if date_from and op.operation_date < date_from:
                continue
            if date_to and op.operation_date > date_to:
                continue

            if op.type.value == "expense":
                expense_data[op.category] = expense_data.get(op.category, 0) + op.amount
            else:
                income_data[op.category] = income_data.get(op.category, 0) + op.amount

        wb = openpyxl.Workbook()

        ws_expense = wb.active
        ws_expense.title = "Расходы по категориям"

        headers = ['Категория', 'Сумма (руб.)']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="e74c3c", end_color="e74c3c", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws_expense.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        row = 2
        total_expense = 0
        for category, amount in sorted(expense_data.items(), key=lambda x: x[1], reverse=True):
            ws_expense.cell(row=row, column=1, value=category)
            ws_expense.cell(row=row, column=2, value=amount)
            total_expense += amount
            row += 1

        ws_expense.cell(row=row, column=1, value="ИТОГО:")
        ws_expense.cell(row=row, column=2, value=total_expense)
        ws_expense.cell(row=row, column=1).font = Font(bold=True)
        ws_expense.cell(row=row, column=2).font = Font(bold=True)

        ws_income = wb.create_sheet("Доходы по категориям")
        for col, header in enumerate(headers, 1):
            cell = ws_income.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="2ecc71", end_color="2ecc71", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        row = 2
        total_income = 0
        for category, amount in sorted(income_data.items(), key=lambda x: x[1], reverse=True):
            ws_income.cell(row=row, column=1, value=category)
            ws_income.cell(row=row, column=2, value=amount)
            total_income += amount
            row += 1

        ws_income.cell(row=row, column=1, value="ИТОГО:")
        ws_income.cell(row=row, column=2, value=total_income)
        ws_income.cell(row=row, column=1).font = Font(bold=True)
        ws_income.cell(row=row, column=2).font = Font(bold=True)

        for col in range(1, 3):
            ws_expense.column_dimensions[get_column_letter(col)].width = 25
            ws_income.column_dimensions[get_column_letter(col)].width = 25

        wb.save(filename)
        messagebox.showinfo("Успех", f"Статистика экспортирована в {filename}")
        return True

    def export_payments(self, filename=None):
        if not filename:
            filename = filedialog.asksaveasfilename(
                title="Экспорт плановых платежей",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not filename:
                return False

        payments = self.app.payment_service.get_payments(self.current_user.id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Плановые платежи"

        headers = ['Название', 'Сумма (руб.)', 'Категория', 'Периодичность', 'Следующий платеж', 'Статус']
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        freq_display = {"daily": "Ежедневно", "weekly": "Еженедельно",
                        "monthly": "Ежемесячно", "yearly": "Ежегодно"}

        for row, payment in enumerate(payments, 2):
            ws.cell(row=row, column=1, value=payment.name)
            ws.cell(row=row, column=2, value=float(payment.amount))
            ws.cell(row=row, column=3, value=payment.category)
            ws.cell(row=row, column=4, value=freq_display.get(payment.frequency.value, payment.frequency.value))
            ws.cell(row=row, column=5,
                    value=payment.next_due_date.strftime('%d.%m.%Y') if payment.next_due_date else "-")
            ws.cell(row=row, column=6, value="Активен" if payment.is_active else "Неактивен")

        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 18

        wb.save(filename)
        messagebox.showinfo("Успех", f"Платежи экспортированы в {filename}")
        return True